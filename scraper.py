import requests
import pandas as pd
from bs4 import BeautifulSoup
import logging
from typing import List, Dict, Optional
import time
import trafilatura
from lxml import html
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import argparse

from config import (
    HEADERS, DEFAULT_CONFIG, MAX_RETRIES, TIMEOUT,
    EXCEL_SETTINGS, TEXT_SETTINGS
)
from utils import setup_logging, extract_article_id, is_advertisement

class VnExpressScraper:
    def __init__(self, url: Optional[str] = None, xpath_config: Optional[Dict] = None):
        setup_logging()
        self.session = requests.Session()
        self.session.headers.update(HEADERS)
        
        # Use custom URL if provided, otherwise use default
        self.url = url if url else DEFAULT_CONFIG['url']
        
        # Use custom XPath config if provided, otherwise use default
        self.xpath_config = xpath_config if xpath_config else DEFAULT_CONFIG['xpath_selectors']

    def fetch_page(self, url: str) -> Optional[str]:
        """Fetch page content with retry logic"""
        for attempt in range(MAX_RETRIES):
            try:
                response = self.session.get(url, timeout=TIMEOUT)
                response.raise_for_status()
                return response.text
            except requests.RequestException as e:
                logging.error(f"Attempt {attempt + 1}/{MAX_RETRIES} failed: {str(e)}")
                if attempt == MAX_RETRIES - 1:
                    logging.error(f"Failed to fetch {url} after {MAX_RETRIES} attempts")
                    return None
                time.sleep(2 ** attempt)  # Exponential backoff
        return None

    def parse_articles(self, html_content: str) -> List[Dict]:
        """Parse article data from HTML content using custom XPath selectors"""
        articles = []
        try:
            # Parse HTML using lxml
            tree = html.fromstring(html_content)
            
            # Find all article elements using custom XPath
            article_elements = tree.xpath(self.xpath_config['article'])
            logging.info(f"Found {len(article_elements)} potential article elements")

            for article in article_elements:
                try:
                    # Extract title and URL using custom XPath
                    title_elements = article.xpath(self.xpath_config['title'])
                    if not title_elements:
                        logging.debug("No title element found")
                        continue

                    title_element = title_elements[0]
                    title = title_element.get(self.xpath_config['title_attr'], '').strip()
                    url = title_element.get(self.xpath_config['url_attr'], '')

                    if not title or not url:
                        logging.debug("Missing title or URL")
                        continue

                    article_id = extract_article_id(url)
                    if not article_id:
                        logging.debug(f"No valid article ID found for URL: {url}")
                        continue

                    # Fetch article content
                    logging.info(f"Fetching content for article: {title}")
                    article_html = self.fetch_page(url)
                    content = ''
                    if article_html:
                        content = trafilatura.extract(article_html, include_comments=False, include_tables=False)
                        if content:
                            logging.info(f"Successfully extracted content for article: {title}")
                        else:
                            logging.warning(f"No content extracted for article: {title}")
                            content = 'Content not available'

                    articles.append({
                        'ID': article_id,
                        'URL': url,
                        'Title': title,
                        'Content': content
                    })
                    logging.info(f"Successfully parsed article: {title}")

                except Exception as e:
                    logging.error(f"Error parsing article: {str(e)}")
                    continue

        except Exception as e:
            logging.error(f"Error parsing HTML content: {str(e)}")
        
        return articles

    def export_to_excel(self, articles: List[Dict]) -> bool:
        """Export articles to Excel file with proper formatting"""
        try:
            if not articles:
                logging.warning("No articles to export")
                return False
                
            # Create DataFrame with only the columns needed for Excel
            excel_data = [{k: v for k, v in article.items() if k != 'Content'} for article in articles]
            df = pd.DataFrame(excel_data)
            
            # Define columns
            columns = EXCEL_SETTINGS['columns']
            
            # Create Excel writer with openpyxl engine
            with pd.ExcelWriter(
                EXCEL_SETTINGS['filename'],
                engine='openpyxl',
                mode='w'
            ) as writer:
                # Convert DataFrame to Excel
                df.to_excel(
                    writer,
                    sheet_name=EXCEL_SETTINGS['sheet_name'],
                    index=False,
                    columns=columns
                )
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets[EXCEL_SETTINGS['sheet_name']]
                
                # Define styles
                header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
                header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Apply header styles
                for col_num, column in enumerate(columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border
                
                # Auto-adjust column widths
                for idx, col in enumerate(columns):
                    column_width = max(
                        len(str(df[col].iloc[i])) for i in range(len(df))
                    )
                    column_width = max(column_width, len(col)) + 2
                    column_letter = get_column_letter(idx + 1)
                    worksheet.column_dimensions[column_letter].width = min(column_width, 50)
                
                # Apply borders and alignment to all cells
                data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                for row in range(2, len(df) + 2):
                    for col in range(1, len(columns) + 1):
                        cell = worksheet.cell(row=row, column=col)
                        cell.border = thin_border
                        cell.alignment = data_alignment
            
            logging.info(f"Successfully exported {len(articles)} articles to {EXCEL_SETTINGS['filename']}")
            return True
            
        except Exception as e:
            logging.error(f"Failed to export to Excel: {str(e)}")
            return False

    def export_to_text(self, articles: List[Dict]) -> bool:
        """Export articles to a readable text file"""
        try:
            if not articles:
                logging.warning("No articles to export to text")
                return False

            logging.info(f"Starting text export of {len(articles)} articles")
            with open(TEXT_SETTINGS['filename'], 'w', encoding=TEXT_SETTINGS['encoding']) as f:
                for i, article in enumerate(articles, 1):
                    logging.debug(f"Writing article {i} to text file")
                    # Write article header
                    f.write(f"Article {i}\n")
                    f.write("-" * 40 + "\n")
                    
                    # Write article metadata
                    f.write(f"ID: {article['ID']}\n")
                    f.write(f"Title: {article['Title']}\n")
                    f.write(f"URL: {article['URL']}\n\n")
                    
                    # Write article content
                    f.write("Content:\n")
                    content = article.get('Content', 'Content not available')
                    f.write(content if content else 'Content not available')
                    
                    # Add separator between articles
                    if i < len(articles):
                        f.write(TEXT_SETTINGS['separator'])

            logging.info(f"Successfully exported {len(articles)} articles to {TEXT_SETTINGS['filename']}")
            return True

        except Exception as e:
            logging.error(f"Failed to export to text file: {str(e)}")
            return False

    def run(self):
        """Main scraping process"""
        logging.info(f"Starting VnExpress scraper for URL: {self.url}")
        
        html_content = self.fetch_page(self.url)
        if not html_content:
            logging.error("Failed to fetch content. Exiting...")
            return

        logging.info("Successfully fetched page content")
        articles = self.parse_articles(html_content)
        logging.info(f"Found {len(articles)} valid articles")

        if articles:
            excel_result = self.export_to_excel(articles)
            text_result = self.export_to_text(articles)
            
            if excel_result and text_result:
                logging.info("Successfully exported articles to both Excel and text formats")
            else:
                logging.error("Failed to export articles to one or both formats")
        else:
            logging.warning("No articles found to export")

def parse_args():
    parser = argparse.ArgumentParser(description='VnExpress Web Scraper')
    parser.add_argument('--url', type=str, help='URL to scrape (default: VnExpress World News)')
    parser.add_argument('--article-xpath', type=str, help='XPath for article elements')
    parser.add_argument('--title-xpath', type=str, help='XPath for title elements')
    parser.add_argument('--title-attr', type=str, help='Attribute name for title')
    parser.add_argument('--url-attr', type=str, help='Attribute name for URL')
    return parser.parse_args()

if __name__ == "__main__":
    args = parse_args()
    
    # Prepare custom XPath config if any arguments are provided
    xpath_config = None
    if any([args.article_xpath, args.title_xpath, args.title_attr, args.url_attr]):
        xpath_config = {
            'article': args.article_xpath or DEFAULT_CONFIG['xpath_selectors']['article'],
            'title': args.title_xpath or DEFAULT_CONFIG['xpath_selectors']['title'],
            'title_attr': args.title_attr or DEFAULT_CONFIG['xpath_selectors']['title_attr'],
            'url_attr': args.url_attr or DEFAULT_CONFIG['xpath_selectors']['url_attr']
        }
    
    scraper = VnExpressScraper(url=args.url, xpath_config=xpath_config)
    scraper.run()
